"""Read formulas from Raw daily data."""
import openpyxl

wb = openpyxl.load_workbook('Trading_Workbook_MASTER.xlsx')
ws = wb['Raw daily data']

print(f"I2: {ws['I2'].value}")
print(f"J2: {ws['J2'].value}")
