"""
Re-merge B31:K31 and B33:K33, center align, keep CF from openpyxl.
"""
import openpyxl
from openpyxl.styles import Alignment, Font

wb = openpyxl.load_workbook('Trading_Workbook_MASTER.xlsx')
ws = wb['Advanced Setup Planner']

# Merge and center B31:K31
print("Merging B31:K31...")
ws.merge_cells('B31:K31')
ws['B31'].alignment = Alignment(horizontal='center', vertical='center')
ws['B31'].font = Font(name='Calibri', size=14, bold=True)

# Merge and center B33:K33
print("Merging B33:K33...")
ws.merge_cells('B33:K33')
ws['B33'].alignment = Alignment(horizontal='center', vertical='center')
ws['B33'].font = Font(name='Calibri', size=12, bold=True)

# Also center B32 (Setup Summary)
ws.merge_cells('B32:K32')
ws['B32'].alignment = Alignment(horizontal='center', vertical='center')

wb.save('Trading_Workbook_MASTER.xlsx')
print("Saved! Open file - LONG/CoreLong should be centered with green bg.")
