"""Fix E column using openpyxl (no auto-fill interference)."""
import openpyxl

wb = openpyxl.load_workbook('Trading_Workbook_MASTER.xlsx')
ws = wb['Daily Log']

# E2: first day = Summary!B10
ws['E2'] = '=IF($A2="","",Summary!$B$10)'

# E3-E41: rolling NAV
for r in range(3, 42):
    ws.cell(row=r, column=5).value = f'=IF($A{r}="","",E{r-1}+H{r-1})'

wb.save('Trading_Workbook_MASTER.xlsx')

# Verify
wb2 = openpyxl.load_workbook('Trading_Workbook_MASTER.xlsx')
ws2 = wb2['Daily Log']
print(f"E2: {ws2['E2'].value}")
print(f"E3: {ws2['E3'].value}")
print(f"E4: {ws2['E4'].value}")
print("Done!")
